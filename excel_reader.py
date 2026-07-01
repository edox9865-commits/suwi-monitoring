# -*- coding: utf-8 -*-
"""DB_표준양식_*.xlsm 파일 읽기 모듈 (수위모니터링 단독 배포용 - 원본 src/data_io/excel_reader.py 발췌)"""

import pandas as pd
import openpyxl
from datetime import datetime


def read_db_file(filepath: str) -> dict:
    """DB_표준양식_*.xlsm 파일에서 측정 자료 읽기. Returns: {'station_name': str, 'data': DataFrame}"""
    wb = openpyxl.load_workbook(filepath, read_only=True, keep_vba=False, data_only=True)

    station_name = ""
    db_sheet = None
    for name in wb.sheetnames:
        try:
            ws = wb[name]
            if not hasattr(ws, "iter_rows"):
                continue
            db_sheet = ws
            break
        except Exception:
            continue

    if db_sheet is None:
        raise ValueError("데이터 시트를 찾을 수 없습니다.")

    all_rows = list(db_sheet.iter_rows(min_row=1, max_row=200, max_col=20, values_only=True))

    _EXCLUDED = {"최소", "최대", "평균", "누계", "합계", "소계", "합", "수위", "유량", "측정", "유속", "면적", "기간"}
    for row in all_rows[:10]:
        if row is None:
            continue
        for cell in row:
            if cell is not None and isinstance(cell, str):
                s = cell.strip()
                if s in _EXCLUDED:
                    continue
                if 2 <= len(s) <= 15 and any(c in s for c in ["교", "댐", "보", "천", "강", "지점"]):
                    if "환산" not in s and "수위유량" not in s:
                        station_name = s
                        break

    dm_header_row = -1
    dm_col = 0
    for i, row in enumerate(all_rows):
        if row is None:
            continue
        for j, cell in enumerate(row):
            if cell is not None and str(cell).strip() == "DM No":
                dm_header_row = i
                dm_col = j
                break
        if dm_header_row >= 0:
            break

    if dm_header_row < 0:
        raise ValueError("'DM No' 헤더를 찾을 수 없습니다. 파일 형식을 확인하세요.")

    header_row = all_rows[dm_header_row]
    col_stage = -1
    col_discharge = -1
    col_datetime = dm_col + 1
    col_instrument = dm_col + 7

    for j, cell in enumerate(header_row):
        if cell is None:
            continue
        s = str(cell).strip().replace("\n", " ")
        if "수위" in s and "H" in s and "수면폭" not in s and "환산" not in s:
            col_stage = j
        elif s == "유량" or (s.startswith("유량") and "구간" not in s):
            col_discharge = j
        elif "일시" in s:
            col_datetime = j
        elif "장비" in s or "유속계" in s:
            col_instrument = j

    if col_stage < 0:
        col_stage = dm_col + 2
    if col_discharge < 0:
        col_discharge = dm_col + 6

    data_start = dm_header_row + 2
    rows_out = []
    for row in all_rows[data_start:]:
        if row is None or all(v is None for v in row):
            continue

        dm_val = row[dm_col] if dm_col < len(row) else None
        if dm_val is None:
            continue
        dm_str = str(dm_val).strip()
        if not dm_str or dm_str in ("None", ""):
            continue

        h_val = _safe_float(row, col_stage)
        q_val = _safe_float(row, col_discharge)

        if h_val is None or q_val is None or q_val <= 0:
            continue
        if h_val > 50 or h_val < -10:
            continue

        dt_raw = row[col_datetime] if col_datetime < len(row) else None
        if isinstance(dt_raw, datetime):
            dt_str = dt_raw.strftime("%Y-%m-%d %H:%M")
        else:
            dt_str = str(dt_raw).strip() if dt_raw else ""

        instrument = ""
        inst_raw = row[col_instrument] if col_instrument < len(row) else None
        if inst_raw is not None:
            instrument = str(inst_raw).strip()

        rows_out.append({
            "dm_no": dm_str,
            "datetime": dt_str,
            "stage": float(h_val),
            "discharge": float(q_val),
            "instrument": instrument,
        })

    wb.close()

    if not rows_out:
        raise ValueError(
            f"측정 자료를 읽지 못했습니다.\n"
            f"헤더행={dm_header_row+1}, 수위컬럼={col_stage+1}, 유량컬럼={col_discharge+1}\n"
            "파일 형식을 확인하거나 수동 입력을 이용하세요."
        )

    return {"station_name": station_name, "data": pd.DataFrame(rows_out)}


def _safe_float(row, col_idx):
    if col_idx < 0 or col_idx >= len(row):
        return None
    val = row[col_idx]
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None
